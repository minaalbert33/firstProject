# imorting openpyxl module to automate excel files.
import openpyxl as pyxl
# imorting time module to use sleep function
import time

# function to make seperate line to format the output.
def seperate_line():
    print(f'---------------------------------------')

# function to load the country file. 
def choice1():
    while True:
        try:
            country = input('Enter the name of the country: ')
            seperate_line()
            # Creating Workbook object of the excel file.
            wb = pyxl.load_workbook(f'{country.lower()}.xlsx')
            global sheet1 
            # Creating Sheet1 object of the first sheet.
            sheet1 = wb['Sheet1']
        except:
            print('This country is not available. ')
            seperate_line()
        else:
            print()
            break
        
# function to print every state and its population and the total population.
def choice2():
    cols2 = list(sheet1.iter_cols())
    # Creating a dictionary, its keys is the states' names.
    # and its values are the states' populations.
    dic2 = dict(zip(cols2[0], cols2[1])) 
    print()
    print('Here is every state and its population: ')
    print()
    total_population = 0
    # printing every state and its population.
    for country, pop in dic2.items():
        print(f'{country.value} --> {pop.value}')
        total_population += pop.value
        seperate_line()
    # printing the total population. 
    print(f'Here is the total population --> {total_population}')
    seperate_line()
    print()
        
# function to print the state with the highest population and the one with the lowest population.
def choice3():
    cols3 = list(sheet1.iter_cols())
    dic3 = dict(zip(cols3[0], cols3[1]))
    
    dict3_values = {}
    # Adding to dict3_values keys that are the cells' values of the first cloumn.
    # Adding to dict3_values values that are the cells' values of the second cloumn.
    for key, value in dic3.items():
        dict3_values[key.value] = value.value
        
    # Creating tuple that contains the state with the highest population and its population
    highest_state = max(zip(dict3_values.keys(), dict3_values.values()))
    print(f'The state with the highest population is: {highest_state[0]} --> {highest_state[1]}')
    seperate_line()
    # Creating tuple that contains the state with the lowest population and its population
    lowest_state = min(zip(dict3_values.keys(), dict3_values.values()))
    print(f'The state with the lowest population is: {lowest_state[0]} --> {lowest_state[1]}')
    seperate_line()
    
# Main App 
def main():
    # printing multiplie lines just to explain the program to user.
    print('Press 1 if you want to load the file of the country')
    seperate_line()
    print('Press 4 if you want to exit from the program.')
    seperate_line()
    print('Canada, Egypt, USA are avalabile.')
    seperate_line()

    # Make user choose to load the file of the country or to exit from the program
    while True:
        try:
            user_input = int(input('Please, Enter Number 1 or 4: '))
            seperate_line()
        except:
            seperate_line()
            print('Invalid Input')
            seperate_line()
            continue
        if user_input != 1 and user_input != 4:
            print('Invalid Number Selection.')
            seperate_line()
            continue
        
        else:
            if user_input == 1:
                choice1()
                break
            else:
                print('See You Soon!')
                exit()
    # printing multiplie lines just to explain the program to user.
    print('Now you have two choices, 2 or 3.')
    seperate_line()
    print('Press 2, if you want to see every state and its population and the total population.')
    seperate_line()
    print('Press 3, if you want to see the state with the highest population or the state with the lowest one.')
    seperate_line()

    while True:
        try:
            user_input = int(input('Great!, Now you should enter 2 or 3: '))
            seperate_line()
        except:
            print('Invalid Input')
            continue
        if user_input != 3 and user_input != 2:
            print('Invalid Number Selection.')
            continue
        else:
            if user_input == 3:
                print('\n')
                choice3()
                exit()
            else:
                choice2()
                exit()

if __name__ == '__main__':
    main()